import streamlit as st
import re
import io
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Control de Calidad de Materiales",
    page_icon="🏗️",
    layout="centered"
)

# ── STYLES ────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title { font-size: 1.6rem; font-weight: 700; color: #1F4E79; margin-bottom: 0.2rem; }
    .sub-title  { font-size: 1rem; color: #5A6A7A; margin-bottom: 1.5rem; }
    .step-box   { background: #F2F7FB; border-left: 4px solid #2E75B6;
                  padding: 0.8rem 1rem; border-radius: 4px; margin-bottom: 0.8rem; }
    .success-box{ background: #E8F5EE; border-left: 4px solid #2E7D5B;
                  padding: 0.8rem 1rem; border-radius: 4px; }
    .warn-box   { background: #FFF8E8; border-left: 4px solid #C27B00;
                  padding: 0.8rem 1rem; border-radius: 4px; }
</style>
""", unsafe_allow_html=True)

# ── HEADER ────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-title">🏗️ Control de Calidad — Recepción de Materiales</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Genera el cuadro de control automáticamente a partir de las mediciones de la obra</div>', unsafe_allow_html=True)
st.markdown("Ref.: **Plantilla RPC Rev. 81 Febrero 2026** · Colegio de Aparejadores de Madrid")
st.divider()

# ── NORMALIZATION ─────────────────────────────────────────────────────────────
SIGLAS = {
    'PVC','PPR','PE','PE100','PEX','EPS','XPS','ACS','BIE','PCI','SATE',
    'ETICS','DEE','CE','UNE','ISO','CTE','RF','SD','HA','HAI',
    'HEB','IPE','HEA','IPN','UPN','ABC','EPDM','HDPE','CHGT','STOC',
    'WR','HR','NF','DIN','PN','DN','EI','EW','REI','RE','DIÁM','NÚM','IMP','IMPERM',
}
SIGLAS_LOWER = {'mm','cm','dm','km','kg','kw','kwp','ml','pa'}
GRAMT = {
    'de','la','el','en','y','a','los','las','del','un','una','con','para',
    'por','se','su','al','o','e','sin','ni','que','es','son','más',
    'pero','sino','como','bien','hay','i','c','s','x','no',
}
RE_CODE = re.compile(r'^(EI\d+\w*|[A-Z]-\d+|\d+[A-Za-z]+-?\d*|[A-Z]\d+|[A-Z]{1,2}-\d+)$')
RE_COMPOUND = re.compile(r'^(\d+[A-Z]+)-(\d+[A-Z]+)$')

def normalize(text):
    if not text: return text
    def protect(m): return m.group(0).replace('-', '\x00')
    t = RE_COMPOUND.sub(protect, text)
    t = re.sub(r'\b([A-Z])-(\d+)', lambda m: m.group(1)+'\x00'+m.group(2), t)
    tokens = re.split(r'(\s+|[/\(\)\[\]",;\+×])', t)
    result = []; first = True
    for tok in tokens:
        if not tok or re.fullmatch(r'[\s/\(\)\[\]",;\+×]+', tok):
            result.append(tok.replace('\x00','-')); continue
        lm = re.match(r'^([^A-Za-z0-9À-ÿ\x00]*)(.+?)([^A-Za-z0-9À-ÿ\x00]*)$', tok)
        if not lm: result.append(tok.replace('\x00','-')); continue
        pre, word, post = lm.group(1), lm.group(2), lm.group(3)
        wo = word.replace('\x00','-'); wu = word.upper().replace('\x00','-'); wl = word.lower().replace('\x00','-')
        if wu.replace('-','') in SIGLAS or wu in SIGLAS:
            result.append(pre+wu+post); first=False; continue
        if wl in SIGLAS_LOWER:
            result.append(pre+wl+post); first=False; continue
        if RE_CODE.match(wo) or '\x00' in word:
            result.append(pre+wo+post); first=False; continue
        if first:
            result.append(pre+wl[0].upper()+wl[1:]+post); first=False; continue
        if wl in GRAMT:
            result.append(pre+wl+post); continue
        result.append(pre+wl+post)
    return ''.join(result).replace('\x00','-')

# ── LOAD RPC TEMPLATE (embedded) ─────────────────────────────────────────────
@st.cache_data
def load_rpc():
    rpc_path = os.path.join(os.path.dirname(__file__), "REV_RPC.xls")
    df = pd.read_excel(rpc_path, sheet_name='RECEPCIÓN PRODUCTOS', header=None)
    skip = ('NOTA','HA ENTRADO','ENTRA EN','NO HA ENTRADO','Además','EN CUMPLIMIENTO',
            'Las Guías','MARCADO','DdP:','CONTENIDO','ETE:','OTROS CONTROLES','Evaluación',
            'ETIQUETADO','(','PRODUCTOS CON','PRODUCTO','Controles y ensayos',
            'Los pavimentos','Los vidrios','Los elementos','EXTINTORES','BOCAS','SISTEMAS')
    products = []; sec = ''
    for i, row in df.iterrows():
        if i < 3: continue
        v = [str(row[j]).strip() if pd.notna(row[j]) else '' for j in range(5)]
        if not v[0] or any(v[0].startswith(p) for p in skip): continue
        if not any(v[1:]):
            sec = v[0]
        else:
            products.append({'sec': sec, 'prod': v[0], 'norma': v[1],
                'ce':  'X' if 'X' in v[2] or 'x' in v[2] else '',
                'ddp': 'X' if 'X' in v[3] or 'x' in v[3] else '',
                'otros': v[4]})
    return products

def g(all_tp, fragment, sec=None):
    for t in all_tp:
        if sec and t['sec'] != sec: continue
        if fragment.lower() in t['prod'].lower(): return t
    return None

def m(prod, norma, ce='', ddp='', otros=''):
    return {'sec':'—','prod':prod,'norma':norma,'ce':ce,'ddp':ddp,'otros':otros}

KNOWN_CHAPTERS = [
    'MOVIMIENTO DE TIERRAS','CIMENTACION','SANEAMIENTO ENTERRADO','ESTRUCTURA',
    'ALBAÑILERIA','FALSOS TECHOS','PAVIMENTOS','REVESTIMIENTOS',
    'IMPERMEABILIZACIONES','AISLAMIENTOS','CARPINTERIA DE MADERA',
    'CARPINTERIA EXTERIOR Y CERRAJERIA','FONTANERIA',
    'ELECTRICIDAD Y TELECOMUNICACIONES','CLIMATIZACION','PCI',
    'VENTILACION GARAJE Y ZZCC','VENTILACION VIVIENDAS','VIDRIO',
    'PINTURA','VARIOS','URBANIZACION','SEGURIDAD Y SALUD',
    'AJARDINAMIENTO','CONTROL DE CALIDAD','GESTION DE RESIDUOS',
]

# ── PARSE PDF ─────────────────────────────────────────────────────────────────
def parse_pdf(pdf_bytes):
    from pdfminer.high_level import extract_text_to_fp
    from pdfminer.layout import LAParams
    laparams = LAParams(line_margin=0.1, word_margin=0.1, char_margin=2.0,
                        boxes_flow=0.5, detect_vertical=False)
    buf = io.StringIO()
    extract_text_to_fp(io.BytesIO(pdf_bytes), buf, laparams=laparams)
    lines = [l.strip() for l in buf.getvalue().split('\n')]

    KNOWN_SET = set(KNOWN_CHAPTERS)
    SKIP_SET  = {'PRESUPUESTO','CÓDIGO','RESUMEN','CANTIDAD',
                 'ALTANA BERROCALES','ALLEGRA BERROCALES'}

    RE_PART_DOT = re.compile(r'^\d{2}\.\d{2}\.\d{2,}[A-Z0-9]*$')
    RE_PART_NUM = re.compile(r'^\d{6,}[A-Z0-9]*$')
    RE_PART_ALT = re.compile(r'^[A-Z][A-Z0-9]{5,11}$')
    RE_SUB      = re.compile(r'^\d{2}\.\d{2}$')
    RE_SUBNUM   = re.compile(r'^\d{4}$')
    RE_CHAP     = re.compile(r'^(0[1-9]|[1-9][0-9])$')
    RE_QTY      = re.compile(r'^[\d.,]+$')
    RE_DATE     = re.compile(r'^\d+ [a-z]+ \d{4}$')
    RE_UNIT     = re.compile(r'^(m2|m3|ml|m |Ud\.? |ud\.? |UD\.? |PA |Kg |kg )')

    def is_code(s):
        if s in KNOWN_SET or s in SKIP_SET: return False
        return bool(RE_PART_DOT.match(s) or RE_PART_NUM.match(s) or RE_PART_ALT.match(s))

    def skip_in_desc(s):
        if not s: return True
        if s in KNOWN_SET or s in SKIP_SET: return True
        if RE_PART_DOT.match(s) or RE_SUB.match(s) or RE_CHAP.match(s): return True
        if RE_PART_NUM.match(s) or RE_SUBNUM.match(s): return True
        if RE_QTY.match(s) or RE_DATE.match(s): return True
        letters = [c for c in s if c.isalpha()]
        if not letters: return True
        has_digits = any(c.isdigit() for c in s)
        upper_ratio = sum(1 for c in letters if c.isupper()) / len(letters)
        return upper_ratio >= 0.95 and not has_digits and not RE_UNIT.match(s)

    def find_desc(start):
        j = start
        while j < len(lines):
            ls = lines[j]
            if skip_in_desc(ls): j += 1; continue
            desc = RE_UNIT.sub('', ls).strip() if RE_UNIT.match(ls) else ls
            if desc and len(desc) > 3 and not RE_QTY.match(desc):
                return desc[:120]
            j += 1
        return None

    chapter_positions = []
    seen_caps = set()
    for i, ls in enumerate(lines):
        if ls in KNOWN_SET and ls not in seen_caps:
            chapter_positions.append((i, ls))
            seen_caps.add(ls)

    def get_chapter(idx):
        cap = ''
        for pos, name in chapter_positions:
            if pos <= idx + 40: cap = name
            else: break
        return cap

    partidas = {}; seen = set()
    for i, ls in enumerate(lines):
        if is_code(ls) and ls not in seen:
            desc = find_desc(i + 1)
            cap  = get_chapter(i)
            if desc and cap:
                seen.add(ls)
                partidas[ls] = {'desc': desc, 'cap': cap}
    return partidas

# ── SYNTHESIS ─────────────────────────────────────────────────────────────────
def synthesize(codes, rpc_prod, partidas_dict):
    unit_re = re.compile(r'^(m2|m3|ml|m\b|Ud\.?\s+|ud\.?\s+|UD\.?\s+|PA\s+|Kg\s+|mes\s+|u\s+|m\s+)', re.I)
    def clean(d): return unit_re.sub('', d).strip()
    def toks(s): return re.split(r'[\s/,]+', s.upper())

    descs = [clean(partidas_dict[c]['desc']) for c in codes if c in partidas_dict]
    if not descs: return normalize(rpc_prod)
    unique = list(dict.fromkeys(descs))
    if len(unique) == 1: return normalize(unique[0])

    tok_lists = [toks(d) for d in unique]
    common = []
    for words in zip(*tok_lists):
        if len(set(words)) == 1: common.append(words[0].title())
        else: break

    if len(common) >= 2:
        prefix = ' '.join(common).rstrip('.,;')
        suffixes = []
        for d in unique:
            tail = ' '.join(toks(d)[len(common):]).rstrip('.,;').lower()
            if tail and tail not in suffixes: suffixes.append(tail)
        if suffixes and len(suffixes) <= 4:
            return normalize((prefix + ' (' + ' / '.join(suffixes) + ')')[:100])
        elif suffixes:
            return normalize((prefix + ' (varios diámetros/tipos)')[:100])
        return normalize(prefix[:100])

    if len(unique) <= 3:
        return normalize((' / '.join(d[:30] for d in unique))[:100])

    rpc_clean = re.sub(
        r'(para aplicaciones en la edificación.*|fabricados en central.*|'
        r'de materias naturales.*|con áridos densos.*|Productos manufacturados.*)',
        '', rpc_prod, flags=re.I).strip().rstrip('.,;-')
    if len(rpc_clean) > 15:
        return normalize((rpc_clean[:70] + ' (uso en obra)')[:100])
    return normalize((unique[0][:60] + ' (y ' + str(len(unique)-1) + ' más)')[:100])

# ── BUILD CHAPTER MAPPING ─────────────────────────────────────────────────────
def build_chapters(all_tp, partidas_dict):
    """Build the chapter→products mapping based on partidas found in the PDF."""
    def G(frag, sec=None): return g(all_tp, frag, sec)

    # Detect which chapters are present in the medición
    caps_present = set(v['cap'] for v in partidas_dict.values() if v['cap'])

    # Full chapter catalogue — same as tested in Allegra/Altana
    # Each entry: (chapter_name, [(product_dict, [codes]), ...])
    # We only include chapters that appear in the PDF
    H  = 'COMPONENTES PARA MORTEROS Y HORMIGONES'
    E  = 'COMPONENTES PARA CIMENTACIÓN Y ESTRUCTURAS  NOTA (11)'
    SAN= 'RED DE SANEAMIENTO'
    ALB= 'ALBAÑILERÍA'
    AIS= 'AISLAMIENTOS TÉRMICOS Y ACUSTICOS'
    FNT= 'INSTALACIONES DE FONTANERÍA, APARATOS SANITARIOS Y GRIFERIAS'

    # Collect partida codes per detected chapter
    def codes_for(cap_fragment):
        return [c for c,v in partidas_dict.items()
                if cap_fragment.upper() in v['cap'].upper()]

    all_known_chapters = [
        'MOVIMIENTO DE TIERRAS','CIMENTACION','SANEAMIENTO ENTERRADO',
        'ESTRUCTURA','ALBAÑILERIA','FALSOS TECHOS','PAVIMENTOS','REVESTIMIENTOS',
        'IMPERMEABILIZACIONES','AISLAMIENTOS','CARPINTERIA DE MADERA',
        'CARPINTERIA EXTERIOR Y CERRAJERIA','FONTANERIA',
        'ELECTRICIDAD Y TELECOMUNICACIONES','CLIMATIZACION','PCI',
        'VENTILACION GARAJE Y ZZCC','VIDRIO','PINTURA','VARIOS','URBANIZACION',
        'SEGURIDAD Y SALUD',
    ]

    # Map each chapter to its RPC products and associated codes
    def chapter_items(cap):
        codes = codes_for(cap)
        if cap == 'MOVIMIENTO DE TIERRAS': return []
        if cap == 'SEGURIDAD Y SALUD': return []

        if cap == 'CIMENTACION':
            return [
                (G('Cementos comunes'), codes),
                (G('Aridos y filleres, de materias naturales, artificiales o reciclados, para hormigones'), codes),
                (G('Aditivos para hormigones en masa'), codes),
                (G('Hormigón fabricado en central'), codes),
                (m('Acero corrugado B-500 SD (armaduras)','CÓDIGO ESTRUCTURAL – Anejo 19 / UNE EN 10080','','','(11)'), codes),
                (G('Fibras de acero para hormigón.'), codes),
                (G('Pilotes de cimentación'), codes),
                (G('Anclajes metálicos para hormigón. Anclajes quimicos'), codes),
                (G('Productos y sistemas para proteger y reparar las estructuras de hormigón, sistemas de prot'), codes),
            ]
        if cap == 'SANEAMIENTO ENTERRADO':
            return [
                (G('Tubos y accesorios de PVC-U (policloruro de vinilo no plastificado) con tubos de pared est'), codes),
                (G('Sumideros sifonicos de PVC'), codes),
                (G('Separadores de grasas'), codes),
                (G('Pozos de registro y camaras de inspección prefabricados de hormigon'), codes),
                (G('Plantas elevadoras de materias fecales'), codes),
                (G('Valvulas de retencion para plantas elevadoras'), codes),
                (G('Canaletas de desagüe de hormigon polimero y PVC'), codes),
                (m('Tubos de drenaje corrugado PE/PVC','UNE EN 13476-1 (CTE)','','','(5) (3*)'), codes),
            ]
        if cap == 'ESTRUCTURA':
            return [
                (G('Hormigón fabricado en central'), codes),
                (G('Cementos comunes'), codes),
                (G('Aridos y filleres, de materias naturales, artificiales o reciclados, para hormigones'), codes),
                (G('Aditivos para hormigones en masa'), codes),
                (m('Acero corrugado B-500 SD (armaduras)','CÓDIGO ESTRUCTURAL – Anejo 19 / UNE EN 10080','','','(11)'), codes),
                (G('Viguetas prefabricadas'), codes),
                (G('Bovedillas de hormigón para Sistemas de forjado'), codes),
                (G('Bovedillas de arcilla cocida'), codes),
                (G('Bovedillas de poliestireno expandido'), codes),
                (G('Condiciones tecnicas de suministro de productos largos'), codes),
                (G('Componentes metálicos para uso en estructuras de acero'), codes),
            ]
        if cap == 'ALBAÑILERIA':
            return [
                (G('Piezas cerámicas para fábrica de albañilería'), codes),
                (G('Bloques de hormigón (con áridos densos y ligeros) para fábrica de albañilería'), codes),
                (G('Morteros de albañilería fabricados en central - Morteros para revoco'), codes),
                (G('Morteros para albañilería fabricados en central, para fábricas'), codes),
                (G('Placas de yeso laminado'), codes),
                (G('Perfilería metálica para particiones'), codes),
                (G('Fijaciones mecanicas para sistemas de placa de yeso laminado'), codes),
                (G('Adhesivos a base de yeso para paneles de yeso'), codes),
                (G('Material para juntas'), codes),
                (G('Angulares y perfiles metálicos para placas de yeso'), codes),
                (G('Materiales en yeso fibroso'), codes),
                (G('Placas de escayola para techos'), codes),
                (G('Sellantes para uso NO estructural en juntas de elementos de fachadas'), codes),
                (m('Placa de cemento para exterior (tipo Aquapanel)','Certificado conformidad fabricante / DdP','','','(3*)'), codes),
            ]
        if cap == 'FALSOS TECHOS':
            return [
                (G('Placas de yeso laminado'), codes),
                (G('Placas de escayola para techos'), codes),
                (G('Perfilería metálica para particiones'), codes),
                (G('Fijaciones mecanicas para sistemas de placa de yeso laminado'), codes),
                (G('Angulares y perfiles metálicos para placas de yeso'), codes),
                (m('Placa de cemento ligera (tipo Aquapanel)','Certificado conformidad fabricante / DdP','','','(3*)'), codes),
            ]
        if cap == 'PAVIMENTOS':
            return [
                (G('Piezas cerámicas para fábrica de albañilería'), codes),
                (G('Piezas de piedra artificial para fábrica de albañilería'), codes),
                (G('Morteros de albañilería fabricados en central - Morteros para revoco'), codes),
                (G('Aridos y filleres, de materias naturales, artificiales o reciclados, para morteros'), codes),
                (G('Cementos comunes'), codes),
                (G('Áridos y polvo mineral, obtenidos de materiales naturales, artificiales o reclicados para'), codes),
                (G('Sellantes en pavimentos de hormigón aplicados en frio'), codes),
                (m('Pavimento vinílico','UNE EN 14041 / DdP fabricante','X','X',''), codes),
            ]
        if cap == 'REVESTIMIENTOS':
            return [
                (G('Morteros de albañilería fabricados en central - Morteros para revoco'), codes),
                (G('Morteros para albañilería fabricados en central, para fábricas'), codes),
                (G('Piezas cerámicas para fábrica de albañilería'), codes),
                (G('Adhesivos a base de yeso para paneles de yeso'), codes),
                (m('Sistema SATE (EPS + mortero + malla refuerzo)','DEE nº 004 (Kit ETICS) / DdP fabricante','','',''), codes),
                (G('Piezas de piedra artificial para fábrica de albañilería'), codes),
            ]
        if cap == 'IMPERMEABILIZACIONES':
            return [
                (G('Membranas líquidas de impermeabilización para uso bajo baldosas'), codes),
                (G('Láminas flexibles anticapilaridad de plástico y caucho'), codes),
                (G('Kits de revestimientos impermeables para suelos y/o paredes de piezas húmedas'), codes),
                (G('Lám. flexibles de plástico y elastómeros'), codes),
            ]
        if cap == 'AISLAMIENTOS':
            return [
                (G('Productos aislantes térmicos para aplicaciones en la edificación. Productos manufacturados', AIS), codes),
                (G('Productos aislantes térmicos para aplicaciones en la edificación. Productos manufacturados'), codes),
                (m('Lámina antimpacto (ruido de impacto bajo pavimento)','DEE 040048 / DEE 040049','','',''), codes),
                (m('Panel multicapa bajo suelo radiante','DdP fabricante / UNE EN 13162 o EN 13164','X','X',''), codes),
                (G('Piezas de hormigón celular curado en autoclave para fábrica de albañilería'), codes),
            ]
        if cap == 'CARPINTERIA DE MADERA':
            return [
                (m('Puerta de madera (paso interior / entrada blindada)','DdP fabricante / Ensayo RF si EI exigido','X','X',''), codes),
                (m('Tablero madera/aglomerado (armarios roperos)','UNE EN 13986 / DdP fabricante','X','X',''), codes),
                (m('Herrajes y accesorios (bisagras, guías correderas)','Certificado del fabricante','','',''), codes),
            ]
        if cap == 'CARPINTERIA EXTERIOR Y CERRAJERIA':
            return [
                (m('Puerta EI2 cortafuego (RF 30 / RF 60)','UNE EN 1634-1 + Certificado ensayo resistencia al fuego','X','X',''), codes),
                (m('Puerta metálica (acceso finca / portal / trastero / garaje)','UNE EN 14351-1 / DdP fabricante','X','X',''), codes),
                (m('Trampilla acceso cubierta','DdP fabricante','','',''), codes),
                (m('Barandilla acero','UNE EN 1090 + DdP fabricante','X','X',''), codes),
                (m('Pasamanos acero inoxidable / lacado','UNE EN 1090 / DdP fabricante','X','X',''), codes),
                (m('Puerta garaje basculante','UNE EN 13241 / DdP fabricante','X','X',''), codes),
                (m('Vierteaguas aluminio lacado','UNE EN 573 / DdP fabricante','X','X',''), codes),
                (m('Carpintería exterior (ventana/balconera, atenuación acústica)','UNE EN 14351-1 / DdP fabricante (declarar Rw)','X','X',''), codes),
            ]
        if cap == 'FONTANERIA':
            return [
                (G('Bañeras de uso doméstico') or G('Bañeras de uso domestico'), codes),
                (G('Platos de ducha de uso domestico'), codes),
                (G('Lavabos'), codes),
                (G('Inodoros y conjuntos de inodoros'), codes),
                (G('Cisternas para inodoros'), codes),
                (G('Mamparas de ducha'), codes),
                (G('Aparatos sanitarios cerámicos'), codes),
                (G('Grifería sanitaria para utilizar en locales'), codes),
                (G('. Griferías sanitarias. Grifos simples'), codes),
                (G('Sellantes para, NO estructural,  para juntas sanitarias'), codes),
                (G('Fregaderos de cocina'), codes),
                (G('Tubos y accesorios de PVC-U (policloruro de vinilo no plastificado) con tubos de pared est'), codes),
                (G('Sumideros sifonicos de PVC'), codes),
                (G('Canaletas de desagüe de hormigon polimero y PVC'), codes),
                (G('Valvulas equilibradoras de presión'), codes),
                (G('instalaciones de agua caliente y fría', FNT), codes),
                (G('Sistemas de canalizacion de polietileno para conduccion de agua para consumo humano'), codes),
                (m('Grupo de presión (abastecimiento agua)','UNE EN 806-3 / DdP fabricante','X','X',''), codes),
                (m('Mueble soporte lavabo','Certificado del fabricante','','',''), codes),
            ]
        if cap == 'ELECTRICIDAD Y TELECOMUNICACIONES':
            return [
                (G('Material de baja tensión'), codes),
                (G('Cables de energía, control y comunicación'), codes),
                (m('Tubo corrugado libre halógenos (canalizaciones eléctricas)','UNE EN 61386 / RD 187/2016','X','',''), codes),
                (G('Sistema generador fotovoltaico'), codes),
                (G('Aparatos para instalaciones de telecomunicaciones'), codes),
            ]
        if cap == 'CLIMATIZACION':
            return [
                (m('Bomba de calor aerotermia','Reglamento UE 813/2013 / DdP fabricante','X','X',''), codes),
                (m('Acumulador / interacumulador ACS','UNE EN 12897 / DdP fabricante','X','X',''), codes),
                (m('Depósito inercia','DdP fabricante','X','X',''), codes),
                (G('instalaciones de agua caliente y fría', FNT), codes),
                (G('Radiadores y convectores'), codes),
                (m('Fancoil (sistema 2 tubos)','Reglamento UE 1253/2014 / DdP fabricante','X','X',''), codes),
                (m('Ascensor eléctrico','RD 88/2013 (Dir. 2014/33/UE) / DdP + marcado CE','X','X',''), codes),
            ]
        if cap == 'PCI':
            return [
                (m('Aljibe PCI + grupo presión contraincendios','RIPCI RD 513/2017 / UNE 23500','','',''), codes),
                (G('Tubos de acero no aleado aptos para soldeo y roscado'), codes),
                (m('BIE 25 mm con armario','UNE EN 671-1 / RIPCI','X','X',''), codes),
                (m('Extintor polvo ABC 6 kg','UNE EN 3-7 + RD 513/2017','X','',''), codes),
                (m('Central detección incendios + detectores + pulsadores','UNE EN 54 (partes aplicables) / RIPCI','X','X',''), codes),
                (m('Central detección CO + detectores CO','UNE EN 50545-1 / RIPCI','X','X',''), codes),
                (m('Sellado cortafuego','UNE EN 1366-3 / DdP fabricante + certificado ensayo RF','X','X',''), codes),
            ]
        if cap == 'VENTILACION GARAJE Y ZZCC':
            return [
                (m('Ventiladores (F300/F200)','Reglamento UE 327/2011 / DdP fabricante','X','X',''), codes),
                (m('Extractores','Reglamento UE 327/2011 / DdP fabricante','X','X',''), codes),
                (G('Compuertas cortafuegos'), codes),
                (m('Conductos ventilación rectangulares','UNE EN 1505 / DdP fabricante','X','X',''), codes),
                (m('Conductos ventilación circulares','UNE EN 1506 / DdP fabricante','X','X',''), codes),
                (m('Central detección CO garaje + detectores CO','UNE EN 50545-1','X','X',''), codes),
            ]
        if cap == 'VIDRIO':
            return [
                (m('Vidrio aislante (doble/triple acristalamiento BE + control solar)','UNE EN 1279-5 / DdP fabricante','X','X',''), codes),
                (m('Vidrio laminado de seguridad (44.2)','UNE EN 14449 / DdP fabricante','X','X',''), codes),
                (m('Espejo canteado / baño','Certificado del fabricante','','',''), codes),
            ]
        if cap == 'PINTURA':
            return [
                (m('Pintura plástica (paramentos interiores)','DdP fabricante / Ficha técnica','','',''), codes),
                (m('Pintura al silicato (fachada exterior)','UNE EN 1062 / DdP fabricante','X','X',''), codes),
                (m('Pintura señalización garaje (suelo)','Ficha técnica del fabricante','','',''), codes),
            ]
        if cap == 'VARIOS':
            return [
                (G('Fregaderos de cocina'), codes),
                (m('Mobiliario de cocina (módulos)','Certificado del fabricante','','',''), codes),
                (m('Horno / vitrocerámica / campana extractora','Directiva 2014/35/UE (BT) + CE','X','',''), codes),
                (m('Encimera de cuarzo','Certificado del fabricante / ficha técnica','','',''), codes),
                (m('Buzón','Certificado del fabricante','','',''), codes),
            ]
        if cap == 'URBANIZACION':
            return [
                (G('Bordillos prefabricados de hormigón'), codes),
                (m('Pavimento elástico zona infantil','UNE EN 1176-1 / UNE EN 1177 / DdP fabricante','X','X',''), codes),
                (m('Solado urbanización (adoquín / losa hormigón)','UNE EN 1338 o EN 1339 / DdP fabricante','X','X',''), codes),
                (G('Geotextiles y productos relacionados, para uso en cimentaciones, movimiento de tierras'), codes),
                (m('Mobiliario urbano','Certificado del fabricante','','',''), codes),
                (m('Juegos infantiles','UNE EN 1176-1 a 7 / UNE EN 1177 / DdP','X','X',''), codes),
                (m('Tierra vegetal / césped / árboles','Albarán + garantía proveedor','','',''), codes),
            ]
        return []

    # Build chapters list preserving order found in PDF
    caps_ordered = list(dict.fromkeys(
        v['cap'] for v in partidas_dict.values() if v['cap']
    ))
    chapters = []
    for cap in caps_ordered:
        items = [(prod, codes) for prod, codes in chapter_items(cap) if prod is not None and codes]
        chapters.append((cap, items))
    return chapters

# ── BUILD EXCEL ───────────────────────────────────────────────────────────────
def build_excel(obra_name, partidas_dict, chapters, all_tp):
    wb = Workbook(); ws = wb.active; ws.title = "Control Recepción Materiales"
    C_HDR="1F4E79"; C_SEC="2E75B6"; C_W="FFFFFF"; C_A="F2F7FB"; C_I="FFF2CC"; CB="9DC3E6"
    thin = Side(style='thin', color=CB)
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    def hf(c): return PatternFill("solid", fgColor=c)
    def cs(ws,r,c,v,bold=False,bg=None,fg="000000",align="left",wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name="Arial",size=9,bold=bold,color=fg)
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
        cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
        cell.border=bord; return cell

    headers=[
        ("Capítulo",22),("Producto / Material",44),("Marca",16),("Modelo",16),
        ("Norma / Especificación",26),("CE\nOblig.",7),("DdP\nOblig.",7),
        ("Otros\ncontroles",12),("Partidas asociadas",24),
        ("Albarán\nrecibido",9),("Marcado CE\nverificado",10),
        ("DdP\nrecibida",9),("Certif./\nOtros",9),
        ("Fecha\nrecepción",11),("Observaciones",30),("✓ Conforme",10),
    ]
    NCOLS=len(headers); LAST=get_column_letter(NCOLS); INPUT_FROM=10

    ws.merge_cells(f"A1:{LAST}1")
    t=ws["A1"]; t.value=f"CUADRO DE CONTROL DE CALIDAD — RECEPCIÓN DE MATERIALES    {obra_name.upper()}"
    t.font=Font(name="Arial",size=12,bold=True,color="FFFFFF"); t.fill=hf(C_HDR)
    t.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=22

    ws.merge_cells(f"A2:{LAST}2")
    s=ws["A2"]; s.value=(f"Ref.: RPC Rev. 81 Febrero 2026 — Colegio Aparejadores de Madrid    "
        f"Obra: {obra_name}    D.E.O.: ____________    Constructora: ____________    Inicio: ____________")
    s.font=Font(name="Arial",size=8,italic=True,color="595959"); s.fill=hf("D6E4F0")
    s.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[2].height=14

    for ci,(h,w) in enumerate(headers,start=1):
        c=ws.cell(row=3,column=ci,value=h)
        c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF"); c.fill=hf(C_HDR)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=bord
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=30; ws.freeze_panes="A4"

    row=4; alt=False
    leyenda_codigos = [
        ('(1)','Sello o Marca de Conformidad a Norma'),
        ('(2)','Certificado del fabricante que acredite la potencia total del equipo de alumbrado'),
        ('(3)','Certificado de homologación o Marca AENOR «N»'),
        ('(3*)','Certificado de conformidad a la norma de aplicación, Orden o Real Decreto'),
        ('(4)','Certificado de homologación por el Ministerio de Industria o DGI autonómica (validez 2 años)'),
        ('(5)','Etiquetado según norma de aplicación, referenciando la misma en el etiquetado o marcado'),
        ('(6)','Etiquetado según norma (placa) y certificado del fabricante del tanque'),
        ('(7)','Informes de ensayos según norma o especificación de aplicación'),
        ('(9)','Certificado del fabricante que acredite la succión en fábricas con categoría de ejecución A'),
        ('(10)','Poliuretano proyectado: marcado, etiquetado e información técnica según UNE 92120-1'),
        ('(11)','Controles y ensayos reglamentarios según CÓDIGO ESTRUCTURAL (art. 4.2.1 y Cap. 5)'),
        ('(12)','Homologación por el Ministerio de Industria o certificación por Organismos de Control'),
        ('(13)','Cementos: certificado de garantía del fabricante s/RC 08 o s/CÓDIGO ESTRUCTURAL'),
        ('(14)','Tipo de elemento estructural y declaración de capacidad portante; marcado según UNE EN 386'),
        ('(X)','Certificado CE de Conformidad. Productos con CE obligatorio por Directiva Europea concreta'),
        ('(X)*','Reacción al fuego: declarar en DdP. Ensayo por lab. acreditado. Plazo máx.: 5 años (RF), 10 años (EF)'),
    ]

    for cap, items in chapters:
        from collections import OrderedDict
        groups = OrderedDict()
        for prod, codes in (items if items else []):
            if not prod or not codes: continue
            desc = synthesize(codes, prod['prod'], partidas_dict)
            if desc not in groups:
                groups[desc]={'normas':[],'ce':prod['ce'],'ddp':prod['ddp'],
                              'otros':set(),'codigos':[],'norma_keys':set()}
            g2=groups[desc]
            n=prod['norma'].strip()
            if n not in g2['norma_keys']: g2['norma_keys'].add(n); g2['normas'].append(n)
            if prod['ce']: g2['ce']=prod['ce']
            if prod['ddp']: g2['ddp']=prod['ddp']
            if prod['otros']: g2['otros'].add(prod['otros'])
            for c2 in codes:
                if c2 not in g2['codigos']: g2['codigos'].append(c2)

        n_items = len(groups)
        ws.merge_cells(f"A{row}:{LAST}{row}")
        c=ws.cell(row=row,column=1,value=f"  ▌ {cap}   ({n_items} material(es) a controlar)")
        c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF"); c.fill=hf(C_SEC)
        c.alignment=Alignment(horizontal="left",vertical="center"); c.border=bord
        ws.row_dimensions[row].height=16; row+=1; alt=False

        if not groups:
            ws.merge_cells(f"A{row}:{LAST}{row}")
            c=ws.cell(row=row,column=1,
                value="  Sin materiales con documentación reglamentaria exigida por la RPC para las partidas de este capítulo.")
            c.font=Font(name="Arial",size=8,italic=True,color="595959"); c.fill=hf("EBF3FB")
            c.alignment=Alignment(horizontal="left",vertical="center"); c.border=bord
            ws.row_dimensions[row].height=14; row+=1; continue

        for desc, g2 in groups.items():
            bg=C_A if alt else C_W
            norma_str='\n'.join(g2['normas']); otros_str=' / '.join(sorted(g2['otros']))
            codigos_str=', '.join(g2['codigos'])
            vals=[cap,desc,"","",norma_str,g2['ce'],g2['ddp'],otros_str,codigos_str,
                  "","","","","","",""]
            for ci,v in enumerate(vals,start=1):
                cs(ws,row,ci,v,bg=C_I if ci>=INPUT_FROM else bg,
                   align="center" if ci in(6,7) else "left",wrap=(ci in(2,5,8,9,15)))
            ws.row_dimensions[row].height=30; row+=1; alt=not alt

    row+=1
    ws.merge_cells(f"A{row}:{LAST}{row}")
    leg=ws.cell(row=row,column=1,value=(
        "LEYENDA  ·  X = Obligatorio exigir  ·  Fondo amarillo = Campo a cumplimentar en obra  "
        "·  CE = Marcado CE  ·  DdP = Declaración de Prestaciones  "
        "·  Los códigos de «Otros controles» se explican en la hoja «Plantilla RPC Completa»  "
        "·  Ref.: Plantilla RPC Rev. 81 Febrero 2026 — Colegio Oficial de Aparejadores de Madrid"))
    leg.font=Font(name="Arial",size=8,italic=True,color="595959"); leg.fill=hf("EBF3FB")
    leg.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); leg.border=bord
    ws.row_dimensions[row].height=24

    # Sheet 2 — RPC + leyenda
    ws2=wb.create_sheet("Plantilla RPC Completa")
    h2=["Sección RPC","Producto / Material","Norma","CE","DdP","Otros controles"]
    w2=[28,58,26,8,8,18]
    for ci,(h,w) in enumerate(zip(h2,w2),start=1):
        c=ws2.cell(row=1,column=ci,value=h)
        c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF"); c.fill=hf(C_HDR); c.border=bord
        c.alignment=Alignment(horizontal="center",vertical="center")
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[1].height=20; ws2.freeze_panes="A2"
    prev_sec=None; r2=2
    for t in all_tp:
        if t['sec']!=prev_sec:
            ws2.merge_cells(f"A{r2}:F{r2}")
            c=ws2.cell(row=r2,column=1,value=f"  {t['sec']}")
            c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF"); c.fill=hf(C_SEC); c.border=bord
            c.alignment=Alignment(horizontal="left",vertical="center")
            ws2.row_dimensions[r2].height=14; r2+=1; prev_sec=t['sec']
        for ci,v in enumerate([t['sec'],t['prod'],t['norma'],t['ce'],t['ddp'],t['otros']],start=1):
            c=ws2.cell(row=r2,column=ci,value=v)
            c.font=Font(name="Arial",size=8)
            c.alignment=Alignment(vertical="center",wrap_text=True); c.border=bord
        ws2.row_dimensions[r2].height=18; r2+=1

    r2+=2
    ws2.merge_cells(f"A{r2}:F{r2}")
    c=ws2.cell(row=r2,column=1,value="  LEYENDA DE CÓDIGOS — COLUMNA «OTROS CONTROLES»")
    c.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); c.fill=hf(C_HDR); c.border=bord
    c.alignment=Alignment(horizontal="left",vertical="center"); ws2.row_dimensions[r2].height=20; r2+=1
    for ci,h in enumerate(["Código","Significado"],start=1):
        c=ws2.cell(row=r2,column=ci,value=h)
        c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF"); c.fill=hf(C_SEC); c.border=bord
        c.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[r2].height=16; r2+=1
    alt2=False
    for codigo,significado in leyenda_codigos:
        bg2=C_A if alt2 else C_W
        c1=ws2.cell(row=r2,column=1,value=codigo)
        c1.font=Font(name="Arial",size=9,bold=True); c1.fill=PatternFill("solid",fgColor=bg2)
        c1.alignment=Alignment(horizontal="center",vertical="center"); c1.border=bord
        ws2.merge_cells(f"B{r2}:F{r2}")
        c2=ws2.cell(row=r2,column=2,value=significado); c2.font=Font(name="Arial",size=9)
        c2.fill=PatternFill("solid",fgColor=bg2)
        c2.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); c2.border=bord
        ws2.row_dimensions[r2].height=28; r2+=1; alt2=not alt2

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── MAIN UI ───────────────────────────────────────────────────────────────────
col1, col2 = st.columns([3, 2])
with col1:
    obra_name = st.text_input("Nombre de la obra", placeholder="Ej: Allegra Berrocales",
                               help="Aparecerá en el título del cuadro Excel")
with col2:
    st.write("")  # spacer

uploaded_file = st.file_uploader(
    "📄 Adjunta el PDF de mediciones de la obra",
    type=["pdf"],
    help="El PDF de presupuesto/mediciones con las partidas agrupadas por capítulo"
)

if uploaded_file and obra_name:
    st.markdown('<div class="step-box">⏳ Procesando mediciones y generando el cuadro…</div>',
                unsafe_allow_html=True)
    try:
        with st.spinner("Leyendo plantilla RPC…"):
            all_tp = load_rpc()

        with st.spinner("Extrayendo partidas del PDF…"):
            pdf_bytes = uploaded_file.read()
            partidas_dict = parse_pdf(pdf_bytes)

        if not partidas_dict:
            st.error("No se han podido extraer partidas del PDF. Comprueba que el archivo sea el presupuesto de mediciones correcto.")
            st.stop()

        caps = list(dict.fromkeys(v['cap'] for v in partidas_dict.values() if v['cap']))
        n_partidas = len(partidas_dict)

        with st.spinner("Cruzando con plantilla RPC y generando Excel…"):
            chapters = build_chapters(all_tp, partidas_dict)
            excel_buf = build_excel(obra_name, partidas_dict, chapters, all_tp)

        n_materials = sum(
            len(set(prod['prod'] for prod, codes in items if prod and codes))
            for _, items in chapters
        )

        st.markdown(f"""
<div class="success-box">
✅ <strong>Cuadro generado correctamente</strong><br>
&nbsp;&nbsp;&nbsp;• <strong>{n_partidas}</strong> partidas procesadas &nbsp;|&nbsp;
<strong>{len(caps)}</strong> capítulos &nbsp;|&nbsp;
<strong>{n_materials}</strong> materiales a controlar
</div>
""", unsafe_allow_html=True)

        fname = f"Control_Recepcion_Materiales_{obra_name.replace(' ','_')}.xlsx"
        st.download_button(
            label="⬇️ Descargar Excel",
            data=excel_buf,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

        with st.expander("Ver capítulos detectados"):
            for cap in caps:
                n = len([c for c,v in partidas_dict.items() if v['cap']==cap])
                st.write(f"**{cap}** — {n} partidas")

    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")
        st.exception(e)

elif uploaded_file and not obra_name:
    st.markdown('<div class="warn-box">⚠️ Introduce el nombre de la obra antes de generar el cuadro.</div>',
                unsafe_allow_html=True)
elif obra_name and not uploaded_file:
    st.markdown('<div class="warn-box">⚠️ Adjunta el PDF de mediciones para continuar.</div>',
                unsafe_allow_html=True)
else:
    st.markdown("""
<div class="step-box">
<strong>Cómo usar esta herramienta:</strong><br>
1. Escribe el nombre de la obra<br>
2. Adjunta el PDF de mediciones (presupuesto sin precios)<br>
3. Descarga el Excel con el cuadro de control listo para obra
</div>
""", unsafe_allow_html=True)

st.divider()
st.caption("Plantilla RPC Rev. 81 · Colegio Oficial de Aparejadores, Arquitectos Técnicos e Ingenieros de Edificación de Madrid · Febrero 2026")
